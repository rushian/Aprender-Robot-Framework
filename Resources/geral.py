"""
Funcoes gerais em python para complementar a acao do Robot
"""

import json
import openpyxl 
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font
import pandas as pd
import platform
import pyscreenshot as CapturaTelaCheia
import random
import xlrd
import docx2txt
from datetime import datetime

def verificar_ambiente():
    """
    Verificar se o python está rodando normalmente, para testar no arquivo geral.py,
    descomente o print(verifica_ambiente)
    """
    versao = platform.python_version()
    sistema = platform.system()
    maquina = platform.machine()
    print('Python executando corretamente')
    print('Versao do python: ', versao)
    print('Sistema: ', sistema)
    print('Maquina: ', maquina)
    numero1 = 5
    numero2 = 3
    print(f'{numero1} + {numero2} = ', str((numero1+numero2)))
    return  versao, sistema, maquina
print(verificar_ambiente())

def ler_qtde_linhas_na_planilha(fileName, sheetname) : 
    workbook = xlrd.open_workbook(fileName)
    worksheet = workbook.sheet_by_name(sheetname)

    rowEndIndex = worksheet.nrows - 1
    colEndIndex = worksheet.ncols - 1 
    print (rowEndIndex)
    return rowEndIndex
#ler_qtde_linhas_na_planilha('C:\\Users\\f675422\\Downloads\\excel (1).xls', 'Sheet0')

def ler_docx(fileName) : 
    my_text = docx2txt.process(fileName)
    print (my_text)
    return my_text
#ler_docx('C:\\Users\\f675422\\Downloads\\teste.docx')

def randomize_value(start: int, end: int):
	"""
	Definir um numero aleatorio no range entre o valor inicial e final
	"""
	value = random.randint(start,end)
	print('\nLimite inicial: ',start,' Limite final: ',end,'\nValor: ',value,flush=True)
	return value
#randomize_value(1,8)

def converter_xls_para_json(caminho_arquivo_xls, caminho_resultado_json):
	"""
	Rotina para gerar json de colunas simples
	"""
	# Carrega o arquivo XLS usando pandas
	data_frame = pd.read_excel(caminho_arquivo_xls) 
	# Converte o DataFrame em uma representação JSON
	dados_pro_json = data_frame.to_json(orient='records') 
    # Obter a data atual
	data_atual = datetime.now().strftime("%d-%m-%Y_%H-%M")
    # Adicionar a data atual ao nome do arquivo
	caminho_resultado_json_com_data = caminho_resultado_json.replace('.json', f'_{data_atual}.json')
	# Converter a string json para um objeto Python
	objeto_json = json.loads(dados_pro_json)
    # Escreve o JSON em um arquivo
	with open(caminho_resultado_json_com_data, 'w') as arquivo_json:
		json.dump(objeto_json,arquivo_json,indent=4)
	print('Arquivo salvo em: ', caminho_resultado_json_com_data)
# Exemplo de uso
# caminho_arquivo_xls = "".join(['TestData', os.sep, 'result1.xlsx'])
# caminho_resultado_json = "".join(['TestData', os.sep, 'massa_simples.json'])
# converter_xls_para_json(caminho_arquivo_xls, caminho_resultado_json)

def converter_xlsx_com_subnivel_para_json(caminho_arquivo_xls, caminho_resultado_json):
    """
    Rotina para gerar json com cabeçalhos encadeados hierarquicamente
    """
    # Carrega o arquivo XLSX usando o openpyxl
    workbook = openpyxl.load_workbook(caminho_arquivo_xls)
    # Seleciona a primeira planilha
    sheet = workbook.active
    # Obtém os valores das células e seus rótulos
    data = []
    for row in sheet.iter_rows(min_row=3,values_only=True):
        data.append(row)
    # Obtém os títulos das colunas
    headers = data.pop(0)
   # Cria um dicionário com os dados
    json_data = {
        "paciente": []
    }
	# Adiciona as linhas da planilha
    for row in data:
        paciente = {
            "dados_pessoais": {
                "nome": row[0],
                "idade": row[1]
            },
            "dados_clinicos": {
                "data_consulta": row[2].strftime("%Y-%m-%d"),
                "medico": row[3],
                "receituario": row[4]
            }
        }
        json_data["paciente"].append(paciente)

    data_atual = datetime.now().strftime("%d-%m-%Y_%H-%M")
    # Adicionar a data atual ao nome do arquivo
    caminho_resultado_json_com_data = caminho_resultado_json.replace('.json', f'_{data_atual}.json')

    # Cria o arquivo JSON com os dados convertidos
    with open(caminho_resultado_json_com_data, 'w') as file:
        json.dump(json_data, file, indent=4)

    print('Arquivo salvo em: ', caminho_resultado_json_com_data)
# Exemplo de uso
# caminho_arquivo_xls = "".join(['TestData', os.sep, 'massa_pacientes.xlsx'])
# caminho_resultado_json = "".join(['TestData', os.sep, 'massa_pacientes.json'])
# converter_xlsx_com_subnivel_para_json(caminho_arquivo_xls, caminho_resultado_json)

def convert_json_to_xlsx(json_file_path, xlsx_file_path):
    """
    Rotina para gerar xlsx a partir do json
    """
    # Carrega o arquivo JSON
    with open(json_file_path, 'r') as file:
        json_data = json.load(file)

    # Extrai os dados do JSON
    pacientes = json_data['paciente']
    dados = []

    # Itera sobre os pacientes e extrai as informações
    for paciente in pacientes:
        nome = paciente['dados_pessoais']['nome']
        idade = paciente['dados_pessoais']['idade']
        data_consulta = paciente['dados_clinicos']['data_consulta']
        medico = paciente['dados_clinicos']['medico']
        receituario = paciente['dados_clinicos']['receituario']

        # Adiciona as informações na lista de dados
        dados.append([nome, idade, data_consulta, medico, receituario])

    # Cria um DataFrame a partir dos dados
    df = pd.DataFrame(dados, columns=['Nome', 'Idade', 'Data da Consulta', 'Médico', 'Receituário'])

    # Cria um novo arquivo XLSX usando openpyxl
    book = Workbook()
    writer = pd.ExcelWriter(xlsx_file_path, engine='openpyxl')
    writer.book = book

    # Obtém a planilha ativa
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    sheet = book.active

    sheet.cell(row=1, column=1).value = 'Paciente'
    sheet.cell(row=2, column=1).value = 'Dados Pessoais'
    sheet.cell(row=2, column=3).value = 'Dados Clinicos'

    # Define a largura das colunas
    for col in sheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Obtém a letra da coluna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Preenche as células com os dados do DataFrame
    for idx, row in df.iterrows():
        for col_idx, value in enumerate(row):
            sheet.cell(row=idx+4, column=col_idx+1).value = value

    # Adiciona os nomes das colunas
    for col_idx, column_name in enumerate(df.columns):
        sheet.cell(row=3, column=col_idx+1).value = column_name


   # Formatação das células
    for row in sheet.iter_rows(min_row=1, max_row=3, min_col=1, max_col=5):
        for cell in row:
            # Define a fonte negrito
            cell.font = Font(bold=True)
            # Centraliza o texto na célula
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # Define as bordas completas
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Ajusta automaticamente a largura das colunas
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
	# Mescla as células nas linhas 1 e 2
    sheet.merge_cells('A1:E1')
    sheet.merge_cells('A2:B2')
    sheet.merge_cells('C2:E2')

    # Salva o arquivo XLSX
    writer.save()

    print('Arquivo salvo em: ', xlsx_file_path)
# Exemplo de uso
# json_file_path = "".join(['TestData', os.sep, 'pacientes.json'])
# xlsx_file_path =  "".join(['TestData', os.sep, 'result1.xlsx'])
# convert_json_to_xlsx(json_file_path, xlsx_file_path)

def captura_tela(dir, nome_arquivo):
	"""
	Captura de tela cheia para obter a barra de tarefas com data e hora 
	* Se a janela atual não for o navegador do teste, o print ira capturar o que estiver na tela
	"""
	imagem = CapturaTelaCheia.grab()
	caminho = dir + nome_arquivo 
	imagem.save(caminho, 'png')
	return 'Evidencia criada em: ' + caminho.replace('\\\\','\\')
#captura_tela('Evidencias\\14-06-2023\\Compra de 3 produtos com sucesso\\','teste1.png')

def pegar_id(elemento):
    """
    Gerar substring do id para usar em click javascript
    """
    inicio = elemento.index('id="')+4
    fim = elemento.index('"]')
    id = elemento[inicio:fim]

    return id
#print(pegar_id('//*[@id="campo-com-valor"]'))

